import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta
import os

# --- CONFIGURATION ET BASE DE DONNÉES ---
DB_NAME = "todo_list.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS tasks
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  gros_titre TEXT,
                  titre TEXT,
                  valeur TEXT,
                  echeance DATE,
                  status TEXT,
                  onglet_origine TEXT)''')
    conn.commit()
    conn.close()

def get_tasks():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query("SELECT * FROM tasks WHERE status = 'En cours' ORDER BY echeance ASC", conn)
    conn.close()
    return df

# --- FONCTION D'IMPORTATION DE TON EXCEL ---
def import_from_excel(file):
    from openpyxl import load_workbook
    wb = load_workbook(file, data_only=True) # data_only=True permet de lire le résultat des formules
    tasks_to_add = []
    
    for sheet_name in wb.sheetnames:
        # On garde ton critère : l'onglet doit commencer par un chiffre
        if sheet_name[0].isdigit(): 
            ws = wb[sheet_name]
            # On récupère le titre en A1
            gros_titre = ws.cell(row=1, column=1).value if ws.cell(row=1, column=1).value else "Sans Titre"
            current_titre = "Général"
            
            for row in range(1, 501):
                cell_v = ws.cell(row=row, column=1)
                val = cell_v.value
                date_val = ws.cell(row=row, column=2).value
                
                if val:
                    # Si c'est en gras -> c'est un Titre (catégorie)
                    if cell_v.font and cell_v.font.bold:
                        current_titre = str(val)
                    else:
                        # Gestion sécurisée de la date
                        try:
                            # On essaie de convertir en date
                            final_date = pd.to_datetime(date_val).date()
                            if pd.isna(final_date): # Si c'est vide ou invalide
                                final_date = datetime.now().date()
                        except:
                            # Si erreur (comme ton =0.6*0.6), on met la date du jour
                            final_date = datetime.now().date()
                        
                        tasks_to_add.append((
                            str(gros_titre), 
                            str(current_titre), 
                            str(val), 
                            final_date,
                            "En cours", 
                            sheet_name
                        ))
    
    if tasks_to_add:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.executemany("INSERT INTO tasks (gros_titre, titre, valeur, echeance, status, onglet_origine) VALUES (?,?,?,?,?,?)", tasks_to_add)
        conn.commit()
        conn.close()
        return True
    return False

# --- INTERFACE STREAMLIT ---
st.set_page_config(page_title="Ma To-Do Intelligente", layout="wide")
init_db()

st.title("🚀 Ma To-Do List Pro")

# Barre latérale pour l'import et les réglages
with st.sidebar:
    st.header("Paramètres")
    uploaded_file = st.file_uploader("Importer l'ancien Excel", type="xlsx")
    if uploaded_file and st.button("Lancer l'importation"):
        import_from_excel(uploaded_file)
        st.success("Données importées !")

    if st.button("Vider la base de données (Reset)"):
        os.remove(DB_NAME)
        st.experimental_rerun()

# --- LOGIQUE DES FILTRES (Tes anciens onglets) ---
tasks_df = get_tasks()
tasks_df['echeance'] = pd.to_datetime(tasks_df['echeance']).dt.date
today = datetime.now().date()

tab_day, tab_week, tab_month, tab_all = st.tabs([
    "📅 Aujourd'hui", "🗓️ Semaine", "📊 Mois", "🗂️ Tout"
])

def display_task_list(df_filtered):
    if df_filtered.empty:
        st.info("Aucune tâche urgente ici.")
    else:
        for i, row in df_filtered.iterrows():
            col1, col2, col3 = st.columns([0.1, 0.7, 0.2])
            with col1:
                if st.checkbox("Fait", key=f"check_{row['id']}"):
                    conn = sqlite3.connect(DB_NAME)
                    conn.execute("UPDATE tasks SET status = 'Terminé' WHERE id = ?", (row['id'],))
                    conn.commit()
                    st.experimental_rerun()
            with col2:
                st.markdown(f"**{row['valeur']}**")
                st.caption(f"{row['gros_titre']} > {row['titre']} | 📍 {row['onglet_origine']}")
            with col3:
                color = "red" if row['echeance'] <= today else "orange"
                st.markdown(f":{color}[{row['echeance']}]")
            st.divider()

with tab_day:
    display_task_list(tasks_df[tasks_df['echeance'] <= today])

with tab_week:
    display_task_list(tasks_df[tasks_df['echeance'] <= today + timedelta(days=7)])

with tab_month:
    display_task_list(tasks_df[tasks_df['echeance'] <= today + timedelta(days=30)])

with tab_all:
    # Formulaire d'ajout rapide
    with st.expander("➕ Ajouter une tâche"):
        with st.form("new_task"):
            f_gt = st.text_input("Gros Titre")
            f_t = st.text_input("Titre")
            f_v = st.text_input("Tâche")
            f_d = st.date_input("Échéance", value=today)
            if st.form_submit_button("Ajouter"):
                conn = sqlite3.connect(DB_NAME)
                conn.execute("INSERT INTO tasks (gros_titre, titre, valeur, echeance, status, onglet_origine) VALUES (?,?,?,?,?,?)",
                             (f_gt, f_t, f_v, f_d, "En cours", "App Mobile"))
                conn.commit()
                st.experimental_rerun()
    
    display_task_list(tasks_df)